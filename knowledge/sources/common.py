#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ADR-013 — shared ingest helpers for the 6 private supplier catalogues.

Resolution is against the FROZEN backbone (never edited). OCR repairs are LOGGED, never silent.
Suspected spelling errors are NOT auto-corrected (routed to the Pass-2 typo path via the delta).
"""
import json, os, re, importlib.util

HERE = os.path.dirname(os.path.abspath(__file__))
K = os.path.join(HERE, "..")
BACKBONE = os.path.join(K, "identity", "botanical_identity.json")
TAXONOMY = os.path.join(K, "taxonomy", "format_codes.json")
CATALOGS = os.path.expanduser("~/Downloads/catalogs")

_spec = importlib.util.spec_from_file_location("bp3", os.path.join(K, "pass3", "build_pass3.py"))
bp3 = importlib.util.module_from_spec(_spec); _spec.loader.exec_module(bp3)
norm = bp3.norm
PART_WORDS = bp3.PART_WORDS
FORM_WORDS = bp3.FORM_WORDS


def nb(s):
    """Normalise non-breaking spaces + collapse whitespace."""
    return re.sub(r"\s+", " ", str(s or "").replace("\xa0", " ")).strip()


def load_backbone():
    bb = json.load(open(BACKBONE, encoding="utf-8"))
    exact, common = bp3.build_indices(bb)
    genera = set()
    for r in bb["identities"]:
        for fld in ("accepted_name", "scientific_synonyms", "original_parsed_names", "trade_synonyms"):
            v = r.get(fld)
            v = [v] if isinstance(v, str) else (v or [])
            for x in v:
                m = re.match(r"^([A-Z][a-z]{2,})\s", x or "")
                if m:
                    genera.add(m.group(1))
    return bb["_meta"]["identity_version"], exact, common, genera


BINOM = re.compile(r"^([A-Z][a-z]+)\s+([a-z]{3,}(?:-[a-z]+)?)")


def ocr_repair(name, genera):
    """Repair OCR-glued Latin names; return (repaired, [ (kind, before, after) ]). Logged, not silent.
    Does NOT fix spelling (typos routed to GBIF later)."""
    fixes = []
    name = nb(name)
    # 1) lowercase->uppercase boundary: TinosporaCordifolia -> Tinospora Cordifolia
    n = re.sub(r"([a-z])([A-Z])", r"\1 \2", name)
    if n != name:
        fixes.append(("case_boundary_split", name, n)); name = n
    # 2) glued all-one-case binomial with a KNOWN genus prefix: Adhatodavasica -> Adhatoda vasica
    if " " not in name and name.isalpha() and len(name) > 8:
        low = name.lower()
        for g in sorted(genera, key=len, reverse=True):
            gl = g.lower()
            if low.startswith(gl) and len(name) - len(gl) >= 3:
                cand = g + " " + name[len(gl):].lower()
                fixes.append(("glued_genus_split", name, cand)); name = cand
                break
    return name, fixes


def extract_latin(text):
    """Return a Genus-species binomial from free text (parens or whole), normalised case, or None."""
    text = nb(text)
    for chunk in re.findall(r"\(([^)]+)\)", text) + [text]:
        m = re.search(r"\b([A-Z][a-zA-Z]+)\s+([A-Za-z][a-z]{2,}(?:-[a-z]+)?)\b", chunk)
        if m:
            g, s = m.group(1), m.group(2)
            if g.lower() not in FORM_WORDS and s.lower() not in FORM_WORDS and s.lower() not in PART_WORDS:
                return g.capitalize() + " " + s.lower()
    return None


def clean_label(label):
    """Trade label -> common-name core candidate (strip form words, parens, OCR noise)."""
    t = nb(label)
    t = re.sub(r"\([^)]*\)", " ", t)
    toks = [w for w in re.split(r"[^A-Za-z]+", t)
            if w and w.lower() not in FORM_WORDS and w.lower() not in PART_WORDS]
    return " ".join(toks).strip()


def resolve(latin, label, exact, common):
    """Resolve strictest-first: accepted->original->scientific->trade (exact) then unique common.
    Returns dict(canonical_id, match_method, matched, [candidates])."""
    cands = []
    if latin:
        cands.append(latin)
    core = clean_label(label)
    if core:
        cands += [core, " ".join(core.split()[:2]), core.split()[0]] if len(core.split()) > 1 else [core]
    for c in cands:
        n = norm(c)
        if n in exact:
            _, method, cid = exact[n]
            return {"canonical_id": cid, "match_method": method, "matched": c}
    uids, amb = set(), None
    for c in cands:
        n = norm(c)
        if n in common:
            ids = common[n]
            if len(ids) == 1:
                uids.add(next(iter(ids)))
            else:
                amb = {"name": c, "candidates": sorted(ids)}
    if len(uids) == 1:
        return {"canonical_id": next(iter(uids)), "match_method": "common_name_exact_unique",
                "matched": core}
    if len(uids) > 1 or amb:
        return {"canonical_id": None, "match_method": "ambiguous",
                "candidates": sorted(uids) if len(uids) > 1 else amb["candidates"]}
    return {"canonical_id": None, "match_method": "unresolved"}


PART = re.compile("|".join(sorted((re.escape(w) for w in PART_WORDS), key=len, reverse=True)), re.I)


def plant_part(*texts):
    for t in texts:
        t = nb(t).lower()
        m = PART.search(t)
        if m:
            return {"raw": m.group(0), "normalized": PART_WORDS[m.group(0).lower()], "confidence": "high"}
    return {"raw": None, "normalized": "unspecified", "confidence": "none"}


def parse_assay_ratio(raw):
    """Parse a mangled assay/ratio cell. ALWAYS keep raw_assay_ratio."""
    r = nb(raw)
    out = {"raw_assay_ratio": raw if raw is not None else "", "ratio": None,
           "assay_percent": None, "marker": None, "method": None}
    if not r:
        return out
    # ratio  a:b  (spreadsheet coerced b to 0b, e.g. 4:01 -> 4:1)
    m = re.search(r"(\d+)\s*:\s*0?(\d+)", r)
    if m:
        out["ratio"] = "%d:%d" % (int(m.group(1)), int(m.group(2)))
    # explicit percent
    mp = re.search(r"(\d+(?:\.\d+)?)\s*%", r)
    if mp:
        out["assay_percent"] = float(mp.group(1))
    elif not m and re.fullmatch(r"0?\.\d+", r):        # decimal fraction: 0.95 -> 95%
        out["assay_percent"] = round(float(r) * 100, 1)
    elif not m and re.fullmatch(r"\d+", r):            # bare integer -> ratio N:1
        out["ratio"] = "%d:1" % int(r)
    # method
    mm = re.search(r"\bby\s+([A-Za-z]+)|\b(HPLC|GC|UV|Titration|Gravimetric|Gr)\b", r, re.I)
    if mm:
        meth = (mm.group(1) or mm.group(2))
        out["method"] = {"gr": "Gravimetric"}.get(meth.lower(), meth.upper() if len(meth) <= 4 else meth.title())
    # marker: capitalised word(s) adjacent to the percent, excluding method words
    if mp:
        seg = r[mp.end():] + " " + r[:mp.start()]
        mk = re.search(r"([A-Z][A-Za-z]+(?:\s+[A-Z]?[A-Za-z]+)?)", seg)
        if mk and mk.group(1).lower() not in ("by", "hplc", "gc", "uv", "titration", "gravimetric"):
            out["marker"] = mk.group(1).strip()
    return out


def map_format(assay, label="", base_default="RE"):
    """-> dict(base_format_code, overlays, standardisation, ratio, normalization_method, review_flags)."""
    lab = nb(label).lower()
    overlays, std, flags, method = [], None, [], "assay_map"
    base = base_default
    if base != "EO":
        if "spray dried" in lab or "spray-dried" in lab:
            base = "SD"
        if assay.get("marker") and assay.get("assay_percent") is not None:
            overlays = ["SE"]                                  # SE overlay never erases the base
            std = {"marker": assay["marker"], "assay": ("%g%%" % assay["assay_percent"]),
                   "method": assay.get("method")}
        if "freeze" in lab:
            if "extract" in lab:                               # freeze-dried EXTRACT: keep base + note
                flags.append("freeze_dried_extract"); method = "manual_review"
            else:                                              # freeze-dried whole material -> FD base
                base = "FD"
    return {"base_format_code": base, "overlays": overlays, "standardisation": std,
            "ratio": assay.get("ratio"), "normalization_method": method, "review_flags": flags}


import csv as _csv  # noqa
from collections import Counter


def process_supplier(meta, raw_rows, exact, common, genera, id_version, out_dir):
    """Resolve + normalise extracted raw rows; write per-supplier outputs; return summary + delta candidates."""
    os.makedirs(out_dir, exist_ok=True)
    rows_out, review, unresolved = [], [], []
    unmapped = Counter()
    mm_counts, fmt_dist, part_dist = Counter(), Counter(), Counter()
    ocr_fixes = []
    base_default = meta.get("base_default", "RE")
    for rr in raw_rows:
        label = nb(rr.get("label"))
        if not label:
            continue
        latin_raw = rr.get("latin")
        fixes = []
        if latin_raw:
            latin, fixes = ocr_repair(latin_raw, genera)
        else:
            latin = extract_latin(label)
        for kind, before, after in fixes:
            ocr_fixes.append({"supplier_id": meta["supplier_id"], "field": "botanical_name",
                              "kind": kind, "before": before, "after": after})
        assay = parse_assay_ratio(rr.get("assay_raw"))
        fmt = map_format(assay, label, base_default)
        res = resolve(latin, label, exact, common)
        mm = res["match_method"]; mm_counts[mm] += 1
        part = plant_part(rr.get("part_text", ""), label)
        sig = fmt["base_format_code"] + ("+" + "+".join(fmt["overlays"]) if fmt["overlays"] else "")
        fmt_dist[sig] += 1; part_dist[part["normalized"]] += 1
        row = {"supplier_product_label": label,
               "raw_botanical_name": latin_raw, "normalized_botanical_name": latin,
               "canonical_id": res["canonical_id"], "match_method": mm,
               "plant_part": part["normalized"], "raw_plant_part": part["raw"],
               "base_format_code": fmt["base_format_code"], "overlays": fmt["overlays"],
               "ratio": fmt["ratio"], "marker": assay["marker"], "assay_percent": assay["assay_percent"],
               "method": assay["method"], "raw_assay_ratio": assay["raw_assay_ratio"],
               "standardisation": fmt["standardisation"], "normalization_method": fmt["normalization_method"],
               "review_flags": fmt["review_flags"], "country_of_origin": rr.get("origin"),
               "ocr_fixes_applied": [f[0] for f in fixes], "source_row": rr.get("source_row")}
        rows_out.append(row)
        if mm == "ambiguous":
            review.append({"type": "ambiguous_common_name", "supplier_id": meta["supplier_id"],
                           "label": label, "candidates": res.get("candidates"), "requires_owner_review": True})
        if fmt["review_flags"]:
            review.append({"type": "form_normalization_disagreement", "supplier_id": meta["supplier_id"],
                           "label": label, "flags": fmt["review_flags"], "requires_owner_review": True})
        if res["canonical_id"] is None and mm == "unresolved":
            unresolved.append({"candidate": latin or clean_label(label),
                               "type": "latin" if latin else "common_only",
                               "supplier_id": meta["supplier_id"], "label": label})

    _dump(os.path.join(out_dir, "%s_rows.json" % meta["supplier_id"]),
          {"_meta": {"supplier_id": meta["supplier_id"], "identity_version": id_version,
                     "source_file": meta["source_file"], "sha256": meta["sha256"],
                     "row_count": len(rows_out)}, "rows": rows_out})
    _dump(os.path.join(out_dir, "%s_review_queue.json" % meta["supplier_id"]),
          {"_meta": {"supplier_id": meta["supplier_id"], "count": len(review)}, "entries": review})
    _dump(os.path.join(out_dir, "unmapped_labels.json"),
          {"_meta": {"supplier_id": meta["supplier_id"]},
           "unmapped": [{"label": k, "count": v} for k, v in unmapped.most_common()]})

    resolved = sum(v for k, v in mm_counts.items() if k not in ("unresolved", "ambiguous"))
    return {"meta": meta, "rows": len(rows_out), "resolved": resolved, "match_methods": dict(mm_counts),
            "ambiguous": mm_counts["ambiguous"], "unresolved": mm_counts["unresolved"],
            "identity_resolution_rate": round(resolved / len(rows_out), 3) if rows_out else 0.0,
            "ocr_fixes": ocr_fixes, "fmt_dist": dict(fmt_dist), "part_dist": dict(part_dist),
            "unresolved_candidates": unresolved}


def _dump(path, obj):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)


def sha256(path):
    import hashlib
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for c in iter(lambda: f.read(1 << 20), b""):
            h.update(c)
    return h.hexdigest()


def load_rows(filename):
    """Return the first sheet as a list-of-lists of nb()-normalised strings."""
    import openpyxl
    p = os.path.join(CATALOGS, filename)
    if not os.path.exists(p):
        raise SystemExit("MISSING CATALOGUE FILE (fail loud): %s" % p)
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[nb(c) if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows, sha256(p), p


def find_cols(header_row, keywords):
    """keyword -> column index (first cell containing the keyword, case-insensitive)."""
    out = {}
    for kw in keywords:
        for i, c in enumerate(header_row):
            if kw.lower() in (c or "").lower():
                out[kw] = i
                break
    return out


def is_num(s):
    return bool(re.fullmatch(r"\d+", (s or "").strip()))
